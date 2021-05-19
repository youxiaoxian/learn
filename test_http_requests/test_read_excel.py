import json
import allure
import requests
from openpyxl import load_workbook
import os

# print(os.path.abspath(__file__))
# print(os.path.dirname(os.path.abspath(__file__)))
file_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "testcase.xlsx")
# print(file_path)
class Excel:
    def __init__(self):
        # 获取工作簿workBook
        self.workbook = load_workbook(file_path)
        # 获取工作表
        self.sheet = self.workbook["Sheet1"]

class ReadExcel(Excel):
    def read_excel(self):
        # 把所有行的数据放到列表中
        test_data = []
        for i in range(2, self.sheet.max_row + 1):
            # 把每行的数据放到字典中
            sub_data = {}
            for j in range(1, self.sheet.max_column + 1):
                sub_data[self.sheet.cell(1, j).value] = self.sheet.cell(i, j).value
            # 拼接每行单元格的数据
            test_data.append(sub_data)
        return test_data

class WriteExcel(Excel):
    def __init__(self):
        super().__init__()
        self.column = GetCol().GetRow()
    def write_title(self):
        self.sheet.cell(1, self.column).value = 'result'
        self.workbook.save(file_path)
    def write_result(self,i,result):
        self.sheet.cell(i+1, self.column).value = result
        self.workbook.save(file_path)

class GetCol(Excel):
    def GetRow(self):
        column = self.sheet.max_column + 1
        return column

@allure.story('正常测试用例')
class TestGwsummaryanalysis:
    def test_Gwana(self):
        cases = ReadExcel().read_excel()
        write=WriteExcel()
        write.write_title()
        for case in cases:
            url = case['host'] + case['path']
            data = json.loads(case['data'])
            httpmethod = case['http_method']
            code = case['expected']
            case_num = case['case_num']
            response = requests.request(method=httpmethod, url=url, json=data)
            assert response.status_code == code
            result=json.dumps(response.json(),indent=2,ensure_ascii=False)
            # result=response.status_code
            write.write_result(case_num,result)